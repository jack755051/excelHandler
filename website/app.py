import re
import os
from flask import Flask, request, render_template, send_file
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# 確認時間格式的正則表達式 =>是否為 11/03/2023 07:44:29 AM
time_pattern = re.compile(r"\d{2}/\d{2}/\d{4}\s\d{2}:\d{2}:\d{2}\s[AP]M")

# 名稱（初始名稱）正則表達式(標準)
name_pattern = re.compile(r"(\d{2,3})[\s_\/-]+(\D+?)(?= \(|$)")

# 名稱（初始名稱）正則表達式(僅數字)
name_pattern_number = re.compile(r"(\d{2,3})(\D+?)(?= \(|$)")

# 名稱（初始名稱）正則表達式 => 可符合
# 003/林煜騰/律師 (123)
# 003/林煜騰/律師
# 017/黃弘宗/穿透式FDA精華液 (穿戴)
name_pattern_new = re.compile(r"(\d{2,3})\/([^\/]+)\/([^\/]+)(?: \(([\d\w]+)\))?")


def format_time(time_str):
    # 將帶日期的時間字符串轉換為不帶日期的時間格式
    return datetime.strptime(time_str, "%m/%d/%Y %I:%M:%S %p").strftime("%I:%M:%S %p")


# 用於判斷是否遲到或早退的函數
def check_lateness(time_str, standard):
    time_obj = datetime.strptime(time_str, "%I:%M:%S %p")
    return time_obj > standard


def check_early_leave(time_str, standard):
    time_obj = datetime.strptime(time_str, "%I:%M:%S %p")
    return time_obj < standard


# 格式化姓名欄位 (新)
def reformat_name(name):
    match_new = name_pattern_new.match(name)
    if match_new:
        # 如果只有两位数字，前面补一个0
        if len(match_new.group(1)) == 2:
            number = "0" + match_new.group(1)
        else:
            number = match_new.group(1)
        # 提取第二个捕获组的内容，并去除首尾空白字符
        name = match_new.group(2).strip()
        # 處理可有可無的()內資料
        if match_new.group(4):
            # 包括括号和括号内的内容
            optional = f" ({match_new.group(4)})"
        else:
            optional = ""
        # 返回格式化后的字符串
        return f"{number}/{name}/{match_new.group(3).strip()}{optional}"
    else:
        print(name)
        return name


# 處理表格大小
def set_column_width(worksheet, width):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 調整欄位寬度
        worksheet.column_dimensions[column_letter].width = (
            adjusted_width if adjusted_width > width else width
        )


# 計算出席率
def calculate_final_results(df):
    # 假設每個人都應該出席
    expected_attendance = 40

    # 計算遲到和早退的人數
    late_count = df["遲到"].str.count("V").sum()
    early_leave_count = df["早退"].str.count("V").sum()

    # 分離出前綴數字，假設它們是以 '/' 分隔的第一個元素
    df["prefix"] = df["名稱（初始名稱）"].str.split("/").str[0]

    # 刪除遲到和早退的記錄
    df = df[~(df["遲到"] == "V") & ~(df["早退"] == "V")]

    # 去重，只保留不重複的前綴數字
    unique_attendance = df["prefix"].drop_duplicates().count()

    # 創建一個新的 DataFrame 來存儲這些值
    final_results_df = pd.DataFrame(
        {
            "應到": [expected_attendance],
            "遲到": [late_count],
            "早退": [early_leave_count],
            "實到": [unique_attendance],
        }
    )

    return final_results_df

    # def simplify_names(df, column_name):
    # 首先，我們創建一個字典來保存每個獨特前綴的最簡潔名稱
    # simplified_name_dict = {}
    # for index, row in df.iterrows():
    #     # 分割 "名稱（初始名稱）" 以獲得前綴和其餘部分
    #     parts = row[column_name].split("/")
    #     prefix = parts[0]
    #     if prefix not in simplified_name_dict:
    #         simplified_name_dict[prefix] = row[column_name]
    #     else:
    #         # 如果當前名稱比已保存的名稱簡潔，則進行替換
    #         if len(row[column_name]) < len(simplified_name_dict[prefix]):
    #             simplified_name_dict[prefix] = row[column_name]

    # # 使用字典更新 DataFrame 中的名稱
    # for index, row in df.iterrows():
    #     prefix = row[column_name].split("/")[0]
    #     df.at[index, column_name] = simplified_name_dict[prefix]

    # return df


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/upload", methods=["GET", "POST"])
def upload():
    # 檢查是否有文件在請求中
    if "file" not in request.files:
        return "NO file part", 400
    file = request.files["file"]
    if file.filename == "":
        return "No selected file", 400
    if file:
        # 確認文件擴展名
        file_ext = file.filename.split(".")[-1].lower()
        # 根據文件擴展名讀取數據
        # 若為 csv 文件
        if file_ext == "csv":
            # 讀取 csv 文件
            df = pd.read_csv(file)
            # 僅留下必要欄位
            df = df[["名稱（初始名稱）", "加入時間", "離開時間"]]
            # 格式化 "名稱（初始名稱）" 欄位
            df["名稱（初始名稱）"] = df["名稱（初始名稱）"].apply(reformat_name)

            # df = simplify_names(df, "名稱（初始名稱）")
            # 定義標準 - 遲到 和 早退
            late_time_standard = datetime.strptime("07:00:00 AM", "%I:%M:%S %p")
            early_leave_standard = datetime.strptime("08:30:00 AM", "%I:%M:%S %p")
            # 處理時間欄位 - 利用 format_time 將 "11/03/2023 07:44:29 AM" 改為 "07:00:00 AM" 格式
            df["加入時間"] = df["加入時間"].apply(
                lambda x: format_time(x) if time_pattern.match(x) else x
            )
            df["離開時間"] = df["離開時間"].apply(
                lambda x: format_time(x) if time_pattern.match(x) else x
            )

            # 去除名稱（初始名稱）中的所有空格
            df["名稱（初始名稱）"] = df["名稱（初始名稱）"].str.replace(r"\s+", "", regex=True)

            # 需要將名稱（初始名稱）欄位整理

            # 篩選名稱欄位，創建兩個不同的sheet - df_with_numbers / df_without_numbers
            df_with_numbers = df[df["名稱（初始名稱）"].str.match(r"\d+.*")]

            # 在篩選後對 df_with_numbers 進行排序
            df_with_numbers["排序號碼"] = (
                df_with_numbers["名稱（初始名稱）"].str.extract(r"(\d+)").astype(int)
            )
            df_with_numbers.sort_values("排序號碼", inplace=True)
            df_with_numbers.drop("排序號碼", axis=1, inplace=True)

            df_without_numbers = df[~df["名稱（初始名稱）"].str.match(r"\d+.*")]

            # df 的聚合操作
            df_grouped = df_with_numbers.groupby("名稱（初始名稱）", as_index=False).agg(
                {"加入時間": "min", "離開時間": "max"}
            )

            # 合并聚合结果回原始 df，这里使用的是外部合并（outer join）
            df_combined = pd.merge(
                df_with_numbers.drop(["加入時間", "離開時間"], axis=1),
                df_grouped,
                on="名稱（初始名稱）",
                how="outer",
            )

            # 现在 df_combined 包含了每个人最早的加入时间和最晚的离开时间
            # 但是，这样会保留所有原始记录，包括重复的 "名稱（初始名稱）"。我们需要删除这些重复记录。

            # 删除除了 '名稱（初始名稱）', '加入時間', '離開時間' 以外的重复项
            df_final = df_combined.drop_duplicates(subset=["名稱（初始名稱）", "加入時間", "離開時間"])

            # 判斷遲到和早退
            df_final["遲到"] = df_final["加入時間"].apply(
                lambda x: "V" if check_lateness(x, late_time_standard) else ""
            )
            df_final["早退"] = df_final["離開時間"].apply(
                lambda x: "V" if check_early_leave(x, early_leave_standard) else ""
            )

            # 過濾不必要的欄位，並新增 "遲到", "早退" 兩欄位
            df_final = df_final[["名稱（初始名稱）", "加入時間", "離開時間", "遲到", "早退"]]

            # 生成最終結果表格
            final_results_df = calculate_final_results(df_final)

            # 轉換 DATAFORM 為EXCEL 文件
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_final.to_excel(writer, sheet_name="遲到早退整理資料", index=False)
                df_without_numbers.to_excel(writer, sheet_name="無法辨識資料", index=False)
                # 新增最終結果的工作表
                final_results_df.to_excel(writer, sheet_name="最終結果", index=False)
                late_time_standard = writer.sheets["遲到早退整理資料"]
                set_column_width(late_time_standard, 15)

            output.seek(0)
            # 設置文件名稱
            # 檔名處理
            base_filename = datetime.now().strftime("%Y%m%d") + "_transform"

            # 判斷版本
            version = 0
            while os.path.exists(
                f"{base_filename}{'' if version == 0 else f'({version})'}.xlsx"
            ):
                version += 1
            excel_file_name = (
                f"{base_filename}{'' if version == 0 else f'({version})'}.xlsx"
            )
            # excel_file_name = file.filename.rsplit(".", 1)[0] + ".xlsx"
            return send_file(
                output,
                download_name=excel_file_name,  # 这里确保参数名正确
                as_attachment=True,  # 这里是 as_attachment 而非 as_attchment
            )
        else:
            return "Unsupported file type", 400


if __name__ == "__main__":
    app.run(debug=True)
